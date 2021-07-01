#!/usr/bin/env python
# -*- coding:utf-8 -*-
# author:GAOXI
# datetime:2020/12/7 12:29
# software: PyCharm

import os
import re
from openpyxl import load_workbook, Workbook


# output = os.popen("pip list").read()
# if "openpyxl" not in output:
#     os.system("pip install -i https://pypi.tuna.tsinghua.edu.cn/simple openpyxl")
#     while True:
#         output = os.popen("pip list").read()
#         if "openpyxl" in output:
#             break



class OperateExcel:
    def __init__(self):

        self.excel_path = r"excel\monkey_result.xlsx"

        self.wb = load_workbook(self.excel_path)
        # self.ws = self.wb.active

    # def get_data(self):
    #     self.ws=self.wb.active
    #     return self.ws

    def write_result(self, sheet_name, row, col, value):
        try:
            self.wb.get_sheet_by_name(sheet_name)
            # self.ws.title = sheet_name
            ws = self.wb[sheet_name]
        except KeyError:
            self.wb.create_sheet(sheet_name, 1)
            ws = self.wb[sheet_name]
        ws.cell(row, col).value = value
        self.wb.save(self.excel_path)
#log的格式：
#ANR in com.google.android.apps.photosgo (com.google.android.apps.photosgo/.home.HomeActivity), time=78679681
#ANR in com.android.chrome (com.android.chrome/org.chromium.chrome.browser.bookmarks.BookmarkActivity)
#CRASH: com.google.android.apps.messaging (pid 929) Time Longth Since Monkey Start: 125718925

class StaticMonkeyLog(OperateExcel):
    def __init__(self, excel_path=None):
        # self.oe=OperateExcel()
        super(StaticMonkeyLog, self).__init__()
        self.packages_list = [] #用于统计所有出现异常的包名,包名可重复[package1,package2,...]
        self.my_list = []  #用于统计package出现的次数[{pckage1:num1},{pckage2:num2}]

    # def read_package_to_list(self,file_path,re_express):
    #     with open(file_path, mode="r") as file:
    #         packages_list = []
    #         line = file.read()
    #         # print(line)
    #         pk_list = re.findall(re_express, line)
    #         for i in range(0, len(pk_list)):
    #             packages_list.append(pk_list[i].replace(",",""))
    #         return packages_list

    def read_package_to_list(self, file_path,re_express): #从monkeylog抓取包名至列表
        with open(file_path, mode="r", encoding="utf8") as file:
            # self.packages_list = []
            while True:
                line = file.readline()
                if not line:
                    break
            # print(line)
                line = re.sub("[,:]", " ", line) #将逗号和冒号替换成空格
                # line = line.replace(",", " ")
                # line = line.replace(":", " ")
                pk_list = re.findall(re_express, line)
                if pk_list:
                    # for i in range(0, len(pk_list)):
                    #     self.packages_list.append(pk_list[i])
                    self.packages_list.append(pk_list[0])
            # print(self.packages_list)
            return self.packages_list

    def make_needinfo(self, packages_list): #汇总信息至mylist中
        # self.my_list = []
        for i in range(0, len(packages_list)):
            is_repeat = False
            for j in range(len(self.my_list)):
                if packages_list[i] == self.my_list[j]["package"]:
                    is_repeat = True
            if not is_repeat:
                my_dict = {}
                appear_count = packages_list.count(packages_list[i])
                my_dict["package"] = packages_list[i]
                my_dict["appear_count"] = appear_count
                self.my_list.append(my_dict)
        # print(self.my_list)
        return self.my_list

    def write_to_excel(self, my_list,sheet_name):  #把mylist（汇总结果）写入excel表中
        self.write_result(sheet_name, row=1, col=1, value="package")
        self.write_result(sheet_name, row=1, col=2, value="appear_count")
        for i in range(0, len(my_list)):
            self.write_result(sheet_name, row=i+2, col=1, value=my_list[i]["package"])
            self.write_result(sheet_name, row=i+2, col=2, value=my_list[i]["appear_count"])

class CollectDdetail(OperateExcel):
    def __init__(self):
        super(CollectDdetail, self).__init__()

    def make_detail(self, file_path, re_express, sheet_name):  #把明细信息写入excel中
        self.write_result(sheet_name, row=1, col=1, value="appear_time")
        self.write_result(sheet_name, row=1, col=2, value="appear_line")
        self.write_result(sheet_name, row=1, col=3, value="package_name")
        with open(file_path, mode="r", encoding="utf8") as file:
            line_count = 0
            excel_row=1
            while True:
                line = file.readline()
                if not line:
                    break
                line_count = line_count+1
                line = re.sub("[,:]", " ", line)  #注意：这里把：和，替换成空格
                pk_list = re.findall(re_express, line)
                if pk_list:
                    appear_time_list = re.findall("[='Start']\s{0,2}(\d+)\\n", line) ###上面已经把：和，替换成空格了
                    if appear_time_list:
                        self.write_result(sheet_name, row=excel_row+1, col=1, value=appear_time_list[0])
                    else:
                        self.write_result(sheet_name, row=excel_row+1, col=1, value="monkeylog中没有时间")
                    self.write_result(sheet_name, row=excel_row+1, col=2, value=line_count)
                    self.write_result(sheet_name, row=excel_row+1, col=3, value=pk_list[0])
                    excel_row = excel_row+1


# if __name__ == '__main__':
    '''
    excel_path = "monkey_result.xlsx"
    if os.path.exists(excel_path):
        os.remove(excel_path)
    wb = Workbook()
    wb.save("excel/monkey_result.xlsx")

    sml = StaticMonkeyLog()
    packages_list = sml.read_package_to_list("monkey_log/monkeylog.txt", "CRASH  (com\S*)")  ###116行把：和，已经替换成空格
    my_list = sml.make_needinfo(packages_list)
    sml.write_to_excel(my_list, "crash_sum")

    sml2 = StaticMonkeyLog()
    packages_list = sml2.read_package_to_list("monkey_log/monkeylog.txt", "ANR in (\S*)")
    my_list = sml2.make_needinfo(packages_list)
    sml2.write_to_excel(my_list, "ANR_sum")

    cd = CollectDdetail()
    cd.make_detail(file_path="monkey_log/monkeylog.txt", re_express="CRASH  (com\S*)",  #\S匹配任何非空白字符
                   sheet_name="crash_detail")  ###上面的代码中把：和，已经替换成空格
    cd.make_detail(file_path="monkey_log/monkeylog.txt", re_express="ANR in (\S*)", sheet_name="ANR_detail")
'''
    # log的格式：
    # ANR in com.google.android.apps.photosgo (com.google.android.apps.photosgo/.home.HomeActivity), time=78679681
    # ANR in com.android.chrome (com.android.chrome/org.chromium.chrome.browser.bookmarks.BookmarkActivity)
    # CRASH: com.google.android.apps.messaging (pid 929) Time Longth Since Monkey Start: 125718925

    # pack = re.findall("CRASH  (com\S*)","CRASH  com.google.android.apps.messaging (pid 929) Time Longth Since Monkey Start  125718925")
    # print(pack)